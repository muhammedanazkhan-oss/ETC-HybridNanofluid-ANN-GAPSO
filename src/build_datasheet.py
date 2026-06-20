"""Build the redesigned thermophysical-property datasheet (multi-sheet, formula-driven)."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT=os.path.join(os.path.dirname(__file__),"..","data","ETC_HybridNanofluid_PropertyDatasheet.xlsx")
HEAD=PatternFill("solid",fgColor="1F4E79"); SUB=PatternFill("solid",fgColor="D9E1F2"); INP=PatternFill("solid",fgColor="FFF2CC")
WHITE=Font(name="Arial",bold=True,color="FFFFFF",size=11); BOLD=Font(name="Arial",bold=True,size=11)
TITLE=Font(name="Arial",bold=True,size=14,color="1F4E79"); REG=Font(name="Arial",size=10); ITAL=Font(name="Arial",italic=True,size=9,color="555555")
thin=Side(style="thin",color="BBBBBB"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center"); LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

def hdr(ws,row,labels,start=1):
    for j,l in enumerate(labels):
        c=ws.cell(row=row,column=start+j,value=l); c.fill=HEAD; c.font=WHITE; c.alignment=CEN; c.border=BORD
def put(ws,row,vals,start=1,bord=True,num=None):
    for j,v in enumerate(vals):
        c=ws.cell(row=row,column=start+j,value=v); c.font=REG; c.alignment=CEN
        if bord: c.border=BORD
        if num and j>=num[0]: c.number_format=num[1]

wb=Workbook()
# ---------------- Cover ----------------
ws=wb.active; ws.title="Cover"
ws["A1"]="Thermophysical-property datasheet"; ws["A1"].font=TITLE
ws["A2"]="ANN-surrogate-driven memetic GA-PSO optimisation of the pumping-power-ratio PEC for a hybrid-nanofluid direct-flow evacuated-tube collector"; ws["A2"].font=BOLD; ws["A2"].alignment=LEFT
ws.merge_cells("A2:H2")
notes=["Author: Muhammed Anaz Khan, University of Bisha, Saudi Arabia",
 "Component 1 = first-named species of each hybrid pair; s1 = mass share of component 1.",
 "Particle properties are treated as constant; carrier properties are temperature-dependent (T in degrees C).",
 "Sheets: Nanoparticles | Carrier fluids (dynamic at T_ref) | Hybrid pairs | Design space | Located optima.",
 "Property values are representative literature data; correlations are those used by the reduced-order solver.",
 "All derived quantities are Excel formulas, so the datasheet recomputes when inputs change."]
for i,n in enumerate(notes): ws.cell(row=4+i,column=1,value=n).font=REG; ws.merge_cells(start_row=4+i,start_column=1,end_row=4+i,end_column=8)
ws.column_dimensions["A"].width=22
for col in "BCDEFGH": ws.column_dimensions[col].width=13

# ---------------- Nanoparticles (Table 3) ----------------
ws=wb.create_sheet("Nanoparticles")
ws["A1"]="Table 3 - Nanoparticle thermophysical properties and morphology"; ws["A1"].font=BOLD
hdr(ws,3,["Material","rho (kg/m3)","cp (J/kg.K)","k (W/m.K)","sphericity psi","shape factor n = 3/psi"])
parts=[["Al2O3",3970,765,40,1.00],["Cu",8933,385,401,1.00],["MWCNT",2100,711,3000,0.30],
       ["Fe3O4",5180,670,9.7,1.00],["Graphene",2250,717,5000,0.20],["TiO2",4250,686,8.9,1.00]]
for i,p in enumerate(parts):
    r=4+i; put(ws,r,p); ws.cell(row=r,column=6,value=f"=3/E{r}").font=REG; ws.cell(row=r,column=6).border=BORD; ws.cell(row=r,column=6).number_format="0.0"; ws.cell(row=r,column=6).alignment=CEN
ws.cell(row=4+len(parts)+1,column=1,value="Sources: representative literature values (see README / manuscript Table 3).").font=ITAL
for col,wd in zip("ABCDEF",[12,13,13,12,13,18]): ws.column_dimensions[col].width=wd

# ---------------- Carrier fluids (Table 4 + dynamic) ----------------
ws=wb.create_sheet("Carrier fluids")
ws["A1"]="Table 4 - Carrier-fluid property correlations and values at a reference temperature"; ws["A1"].font=BOLD
ws["A3"]="Reference temperature T_ref (deg C):"; ws["A3"].font=BOLD; ws.merge_cells("A3:C3")
ws["D3"]=25; ws["D3"].fill=INP; ws["D3"].font=Font(name="Arial",bold=True,color="0000FF"); ws["D3"].border=BORD; ws["D3"].alignment=CEN
ws["E3"]="<- editable input (blue); all values below recompute"; ws["E3"].font=ITAL; ws.merge_cells("E3:K3")
# coefficient table: rho=a+bT+cT2 ; k=a+bT+cT2 ; cp=a+bT ; mu form per carrier
hdr(ws,5,["Carrier","a_rho","b_rho","c_rho","a_k","b_k","c_k","a_cp","b_cp","mu model","mu p1","mu p2","mu p3"])
coef=[
 ["Distilled water",1000.6,-0.0128,-0.0035,0.5582,0.00214,-9.5e-6,4180,0,"A*10^(B/(T+C))",2.414e-5,247.8,133.15],
 ["EG/water 60:40",1093,-0.55,0,0.380,0.0005,0,3280,3.0,"a*EXP(b*T)",6.93e-3,-0.0216,""],
 ["Synthetic HTF oil",900,-0.70,0,0.137,-8e-5,0,1800,3.5,"a*EXP(b*T)",5.77e-2,-0.0339,""]]
for i,c in enumerate(coef):
    r=6+i; put(ws,r,c)
    for col in range(2,14): ws.cell(row=r,column=col).number_format="General"
# computed properties at T_ref
ws.cell(row=10,column=1,value="Computed at T_ref:").font=BOLD
hdr(ws,11,["Carrier","rho (kg/m3)","cp (J/kg.K)","k (W/m.K)","mu (Pa.s)","mu (mPa.s)"])
for i in range(3):
    src=6+i; r=12+i
    name=coef[i][0]
    ws.cell(row=r,column=1,value=name).font=REG; ws.cell(row=r,column=1).border=BORD; ws.cell(row=r,column=1).alignment=CEN
    Tr="$D$3"
    ws.cell(row=r,column=2,value=f"=B{src}+C{src}*{Tr}+D{src}*{Tr}^2")
    ws.cell(row=r,column=3,value=f"=H{src}+I{src}*{Tr}")
    ws.cell(row=r,column=4,value=f"=E{src}+F{src}*{Tr}+G{src}*{Tr}^2")
    if i==0:  # water Vogel
        ws.cell(row=r,column=5,value=f"=K{src}*10^(L{src}/({Tr}+M{src}))")
    else:     # exp form
        ws.cell(row=r,column=5,value=f"=K{src}*EXP(L{src}*{Tr})")
    ws.cell(row=r,column=6,value=f"=E{r}*1000")
    for col in range(2,7):
        cc=ws.cell(row=r,column=col); cc.font=REG; cc.border=BORD; cc.alignment=CEN
        cc.number_format="0.000" if col in(4,5,6) else "0.0"
    ws.cell(row=r,column=5).number_format="0.000000"
ws.cell(row=16,column=1,value="rho,k = a+bT+cT^2 ; cp = a+bT ; water mu Vogel A*10^(B/(T+C)); EG/water & oil mu = a*exp(b*T).").font=ITAL
ws.merge_cells("A16:M16")
for col,wd in zip("ABCDEFGHIJKLM",[16,9,9,10,9,9,10,8,8,14,10,9,9]): ws.column_dimensions[col].width=wd

# ---------------- Hybrid pairs (Table 1) ----------------
ws=wb.create_sheet("Hybrid pairs")
ws["A1"]="Table 1 - Hybrid-pair definitions"; ws["A1"].font=BOLD
hdr(ws,3,["Hybrid pair","Component 1 (first-named)","Component 2","Higher-k component"])
pp=[["Al2O3-Cu","Al2O3 (k=40)","Cu (k=401)","Cu (component 2)"],
    ["MWCNT-Fe3O4","MWCNT (k=3000)","Fe3O4 (k=9.7)","MWCNT (component 1)"],
    ["Graphene-TiO2","Graphene (k=5000)","TiO2 (k=8.9)","Graphene (component 1)"]]
for i,p in enumerate(pp): put(ws,4+i,p)
ws.cell(row=8,column=1,value="Mass split: s1 = w1/w (component-1 mass share); w1 = s1*w, w2 = (1-s1)*w.").font=ITAL
for col,wd in zip("ABCD",[16,24,16,22]): ws.column_dimensions[col].width=wd

# ---------------- Design space (Table 2) ----------------
ws=wb.create_sheet("Design space")
ws["A1"]="Table 2 - Full-factorial design space of the controlled inputs"; ws["A1"].font=BOLD
hdr(ws,3,["Variable","Levels","Count"])
ds=[["Hybrid pair","Al2O3-Cu; MWCNT-Fe3O4; Graphene-TiO2",3],
 ["Carrier fluid","Distilled water; EG/water 60:40; Synthetic HTF oil",3],
 ["Total weight fraction w (%)","0.25, 0.5, 1.0, 1.5, 2.0, 2.5, 3.0",7],
 ["Component-1 mass share 100*s1 (%)","25, 50, 75",3],
 ["Volumetric flow rate Vdot (L/min)","0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 6.0",8],
 ["Inlet temperature Ti (deg C)","20, 40, 60, 80",4],
 ["Irradiance Is (W/m2)","200, 700, 1200",3],
 ["Ambient temperature Ta (deg C)","15, 25, 35",3]]
for i,d in enumerate(ds):
    r=4+i; ws.cell(row=r,column=1,value=d[0]); ws.cell(row=r,column=2,value=d[1]); ws.cell(row=r,column=3,value=d[2])
    for col in (1,2,3):
        cc=ws.cell(row=r,column=col); cc.font=REG; cc.border=BORD; cc.alignment=LEFT if col==2 else CEN
ws.cell(row=12,column=1,value="Total runs ="); ws.cell(row=12,column=1).font=BOLD
ws.cell(row=12,column=3,value="=PRODUCT(C4:C11)"); ws.cell(row=12,column=3).font=BOLD; ws.cell(row=12,column=3).border=BORD; ws.cell(row=12,column=3).alignment=CEN
for col,wd in zip("ABC",[30,40,8]): ws.column_dimensions[col].width=wd

# ---------------- Located optima (Table 11) ----------------
ws=wb.create_sheet("Located optima")
ws["A1"]="Table 11 - Directly re-optimised pumping-power-ratio optima (manuscript-reported)"; ws["A1"].font=BOLD
hdr(ws,3,["Pair","Carrier","w* (%)","100*s1 (%)","Vdot* (L/min)","phi_hnf","dP (kPa)","PEC_f","PEC_eps","dPEC = PEC_eps-1"])
opt=[["Al2O3-Cu","Distilled water",0.56,25,0.50,0.0008,0.13,1.0016,0.999997],
 ["Al2O3-Cu","EG/water 60:40",3.00,25,1.71,0.0048,1.58,1.0085,0.999909],
 ["Al2O3-Cu","Synthetic HTF oil",0.25,25,0.50,0.0003,0.60,1.0003,0.999570],
 ["Graphene-TiO2","Distilled water",0.25,25,0.50,0.0007,0.13,0.9992,0.998602],
 ["Graphene-TiO2","EG/water 60:40",0.25,25,0.50,0.0008,0.22,0.9992,0.998619],
 ["Graphene-TiO2","Synthetic HTF oil",0.25,25,0.50,0.0006,0.61,0.9996,0.998969],
 ["MWCNT-Fe3O4","Distilled water",0.25,25,0.50,0.0006,0.13,0.9994,0.998782],
 ["MWCNT-Fe3O4","EG/water 60:40",0.25,25,1.62,0.0007,1.24,0.9994,0.998799],
 ["MWCNT-Fe3O4","Synthetic HTF oil",0.25,25,0.50,0.0006,0.61,0.9997,0.999052]]
for i,o in enumerate(opt):
    r=4+i; put(ws,r,o)
    ws.cell(row=r,column=10,value=f"=I{r}-1"); ws.cell(row=r,column=10).border=BORD; ws.cell(row=r,column=10).number_format="0.0E+00"; ws.cell(row=r,column=10).alignment=CEN; ws.cell(row=r,column=10).font=REG
    ws.cell(row=r,column=8).number_format="0.0000"; ws.cell(row=r,column=9).number_format="0.000000"; ws.cell(row=r,column=6).number_format="0.0000"
ws.cell(row=14,column=1,value="Under PEC_eps no combination exceeds parity (all dPEC < 0). Best near-parity case: Al2O3-Cu / EG-water.").font=ITAL
ws.merge_cells("A14:J14")
for col,wd in zip("ABCDEFGHIJ",[14,17,8,11,13,9,9,9,11,16]): ws.column_dimensions[col].width=wd

wb.save(OUT); print("saved", os.path.relpath(OUT))
